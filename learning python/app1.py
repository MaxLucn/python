# def lbs_to_kg(weight):
#     return weight * 0.45
#
#
# def kg_to_lbs(weight):
#     return weight / 0.45

# 自己建议一个文件定义类、函数，然后在这里将它作为一个模块引入，更清楚的知道模块引入的方式
# import converters
# from converters import kg_to_lbs
#
# kg_to_lbs(60)
#
# print(converters.kg_to_lbs(79))


# from utils import find_min
#
# numbers = [2, 5, 6, 8, 0, 34, 81]
# my_min = find_min(numbers)
# print(my_min)

# from ecommerce.shipping import shipping
#
#
# shipping.calc_shipping()


#  生成随机值模块的演示，这是一个 python3 中已经建立好的标准库，
# import random
#
# for i in range(3):
#    print(random.random())

#  随机值的练习
# import random
#
#
# class Dice:
#     def roll(self):
#         first = random.randint(2, 9)
#         second = random.randint(1, 7)
#         return first, second
#
#
# dice = Dice()
# print(dice.roll())


# 在当前目录下的。py文件
# from pathlib import Path
#
#
# path = Path()
# for file in path.glob("*.xlsx"):
#     print(file)



# 自动化的测试，让表格的的值自动填充，并且绘出相应的直方图
# 存在的问题：无法打开  .xlsx表格，错误代码：
# KeyError: "There is no item named '[Content_Types].xml' in the archive"
#

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# 打开这个表格，找到想要改变的值的地方
wb = xl.load_workbook(" test automation.xlsx")
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
# print(sheet.max_row)

# 确保要改变的值的位置以及跟它有关系的值的位置，并借助已经存在的值跟想要的值的关系自动填充值
for row in range(2, sheet.max_row + 1):
    # print(row)
    cell = sheet.cell(row, 2)
    corrected_price = cell.value * 0.14
    corrected_price_cell = sheet.cell(row, 3)
    corrected_price_cell.value = corrected_price

# 绘出对应的直方图，并给定位置
Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
wb.save('test automation2.xlsx')
# 现在这个程序只会修改特定的表格，可以作出以下的改变
# def pross_workbook(filename):
#     wb = xl.load_workbook(filename)
#     sheet = wb['Sheet1']
#
#     ...
#
#
# ...
#
#     wb.save(filename)
# 这样的好处是这个程序不止只能改变这一个表格，只要把需要改变的表格名传递给 filename 就可以改变任何需要
# 作出这种改变的表格