# def calc_shipping():
#     print("calc_shipping")
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# 打开这个表格，找到想要改变的值的地方
def pross_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # print(sheet.max_row)

    # 确保要改变的值的位置以及跟它有关系的值的位置，并借助已经存在的值跟想要的值的关系自动填充值
    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 2)
        corrected_price = cell.value * 0.14
        corrected_price_cell = sheet.cell(row, 3)
        corrected_price_cell.value = corrected_price

    # 绘出对应的直方图，并给定位置
    Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)